"""
Instagram DM Lead Manager
A simple CRM for small businesses tracking leads from Instagram DMs.
Run with: streamlit run streamlit_app.py
"""

import streamlit as st
import sqlite3
import pandas as pd
import io
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# DATABASE SETUP
# ─────────────────────────────────────────────

DB_PATH = "leads.db"

def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_connection() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS leads (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                name          TEXT NOT NULL,
                username      TEXT NOT NULL,
                phone         TEXT,
                inquiry       TEXT,
                status        TEXT DEFAULT 'New',
                notes         TEXT,
                date_added    TEXT NOT NULL,
                followup_date TEXT
            )
        """)
        conn.commit()


# ─────────────────────────────────────────────
# CRUD OPERATIONS
# ─────────────────────────────────────────────

def add_lead(name, username, phone, inquiry, status, notes, followup_date):
    today = date.today().isoformat()
    followup_str = followup_date.isoformat() if followup_date else None
    with get_connection() as conn:
        conn.execute("""
            INSERT INTO leads (name, username, phone, inquiry, status, notes, date_added, followup_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, username, phone, inquiry, status, notes, today, followup_str))
        conn.commit()


def get_all_leads():
    with get_connection() as conn:
        df = pd.read_sql_query("SELECT * FROM leads ORDER BY date_added DESC", conn)
    return df


def update_lead(lead_id, status, notes, followup_date):
    followup_str = followup_date.isoformat() if followup_date else None
    with get_connection() as conn:
        conn.execute("""
            UPDATE leads SET status = ?, notes = ?, followup_date = ?
            WHERE id = ?
        """, (status, notes, followup_str, lead_id))
        conn.commit()


def delete_lead(lead_id):
    with get_connection() as conn:
        conn.execute("DELETE FROM leads WHERE id = ?", (lead_id,))
        conn.commit()


def insert_demo_data():
    today = date.today()
    samples = [
        ("Aisha Mehta",    "@aisha.mehta",     "9876543210", "Interested in custom embroidery hoodies",       "Interested",      "Wants 5 pieces in navy blue. Waiting for size chart.", (today + timedelta(days=2)).isoformat()),
        ("Rohan Kapoor",   "@rohankapoor_",    "",           "Asked for pricing on logo tote bags",            "Follow-up",       "Sent price list. No reply yet.",                      (today - timedelta(days=1)).isoformat()),
        ("Sneha D",        "@sneha.deshpande", "9123456789", "Wants birthday gift hamper, budget Rs.1500",     "New",             "",                                                    None),
        ("Vikram Nair",    "@vikram_nair22",   "",           "Bulk order inquiry - 50 custom mugs for office", "Closed",          "Order confirmed. Payment received.",                  None),
        ("Priya Sharma",   "@priyasharma_art", "8800001234", "Wants to know if we ship to Pune",              "Not Interested",  "Said shipping cost was too high.",                    None),
        ("Kabir Hussain",  "@kabirhussain",    "",           "Custom phone case with photo print",             "Interested",      "Shared 3 reference images. Awaiting final approval.", (today + timedelta(days=5)).isoformat()),
        ("Meera Joshi",    "@meeraj_99",       "9988776655", "Wedding return gifts - 100 qty, personalised",  "Follow-up",       "Big potential order. Call scheduled.",                today.isoformat()),
    ]
    with get_connection() as conn:
        for i, s in enumerate(samples):
            conn.execute("""
                INSERT INTO leads (name, username, phone, inquiry, status, notes, date_added, followup_date)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (*s[:6], (today - timedelta(days=len(samples) - i)).isoformat(), s[6]))
        conn.commit()


# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────

# Status color map: (background hex, text hex) — no # prefix for openpyxl
STATUS_FILL = {
    "New":           ("DBEAFE", "1D4ED8"),
    "Interested":    ("DCFCE7", "166534"),
    "Not Interested":("FEE2E2", "991B1B"),
    "Follow-up":     ("FEF9C3", "854D0E"),
    "Closed":        ("F3F4F6", "374151"),
}

def _border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def generate_excel(df: pd.DataFrame) -> bytes:
    """Build a formatted 3-sheet .xlsx from leads data and return as bytes."""
    wb = Workbook()

    # ════════════════════════════════════════════
    # SHEET 1 — All Leads
    # ════════════════════════════════════════════
    ws = wb.active
    ws.title = "All Leads"

    # Title
    ws.merge_cells("A1:I1")
    tc = ws["A1"]
    tc.value = f"InstaLeads Export  —  {date.today().strftime('%d %b %Y')}"
    tc.font      = Font(name="Arial", bold=True, size=13, color="0F0F0F")
    tc.fill      = PatternFill("solid", fgColor="F8F8F8")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ["#", "Name", "Instagram", "Phone", "Inquiry", "Status", "Notes", "Date Added", "Follow-up Date"]
    h_fill  = PatternFill("solid", fgColor="0F0F0F")
    h_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = h_align; c.border = _border()
    ws.row_dimensions[2].height = 26

    # Column widths
    for ci, w in enumerate([5, 20, 20, 16, 42, 16, 36, 14, 16], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        status = row.get("status", "New")
        bg, fg = STATUS_FILL.get(status, ("FFFFFF", "000000"))
        values = [
            ri - 2,
            row.get("name", ""),
            row.get("username", ""),
            row.get("phone", ""),
            row.get("inquiry", ""),
            status,
            row.get("notes", ""),
            row.get("date_added", ""),
            row.get("followup_date", ""),
        ]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _border()
            if ci == 6:  # Status
                c.font      = Font(name="Arial", size=9, bold=True, color=fg)
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 1:  # Row #
                c.font      = Font(name="Arial", size=9, color="9CA3AF")
                c.alignment = Alignment(horizontal="center", vertical="top")
            else:
                c.font      = Font(name="Arial", size=9)
                c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[ri].height = 40

    ws.freeze_panes = "A3"

    # ════════════════════════════════════════════
    # SHEET 2 — Summary
    # ════════════════════════════════════════════
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14

    ws2.merge_cells("A1:B1")
    t = ws2["A1"]
    t.value = "Lead Summary"
    t.font = Font(name="Arial", bold=True, size=13); t.fill = PatternFill("solid", fgColor="F8F8F8")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 30

    for ci, lbl in [(1, "Status"), (2, "Count")]:
        c = ws2.cell(row=2, column=ci, value=lbl)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F0F0F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    ws2.row_dimensions[2].height = 24

    statuses = ["New", "Interested", "Not Interested", "Follow-up", "Closed"]
    for i, s in enumerate(statuses, start=3):
        cnt = len(df[df["status"] == s]) if not df.empty else 0
        bg, fg = STATUS_FILL.get(s, ("FFFFFF", "000000"))

        cs = ws2.cell(row=i, column=1, value=s)
        cs.font = Font(name="Arial", bold=True, size=9, color=fg)
        cs.fill = PatternFill("solid", fgColor=bg)
        cs.alignment = Alignment(horizontal="left", vertical="center")
        cs.border = _border()

        cc = ws2.cell(row=i, column=2, value=cnt)
        cc.font = Font(name="Arial", size=10, bold=True)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border = _border()
        ws2.row_dimensions[i].height = 22

    # Total row
    tr = len(statuses) + 3
    for ci, val in [(1, "TOTAL"), (2, f"=SUM(B3:B{tr-1})")]:
        c = ws2.cell(row=tr, column=ci, value=val)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F0F0F")
        c.alignment = Alignment(horizontal="center" if ci == 2 else "left", vertical="center")
        c.border = _border()
    ws2.row_dimensions[tr].height = 24

    # ════════════════════════════════════════════
    # SHEET 3 — Follow-ups
    # ════════════════════════════════════════════
    ws3 = wb.create_sheet("Follow-ups")
    for ci, w in zip([1,2,3,4,5], [20, 20, 16, 16, 20]):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    ws3.merge_cells("A1:E1")
    t3 = ws3["A1"]
    t3.value = "Follow-up Tracker"
    t3.font = Font(name="Arial", bold=True, size=13)
    t3.fill = PatternFill("solid", fgColor="F8F8F8")
    t3.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 30

    fu_headers = ["Name", "Username", "Status", "Follow-up Date", "Status Label"]
    for ci, h in enumerate(fu_headers, 1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F0F0F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    ws3.row_dimensions[2].height = 26

    if not df.empty:
        fu_df = df[df["followup_date"].notna() & (df["followup_date"] != "")].copy()
        fu_df["_days"] = pd.to_datetime(fu_df["followup_date"], errors="coerce").dt.date.apply(
            lambda d: (d - date.today()).days if d else None
        )
        fu_df = fu_df.sort_values("_days")

        for ri, (_, row) in enumerate(fu_df.iterrows(), start=3):
            days = row.get("_days")
            if days is None:
                label, row_bg = "—", "FFFFFF"
            elif days < 0:
                label, row_bg = f"{abs(int(days))} days overdue", "FEE2E2"
            elif days == 0:
                label, row_bg = "Due today", "FEF9C3"
            else:
                label, row_bg = f"In {int(days)} days", "DCFCE7"

            fill = PatternFill("solid", fgColor=row_bg)
            for ci, val in enumerate([
                row.get("name",""), row.get("username",""),
                row.get("status",""), row.get("followup_date",""), label
            ], 1):
                c = ws3.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", size=9)
                c.fill = fill
                c.alignment = Alignment(vertical="center")
                c.border = _border()
            ws3.row_dimensions[ri].height = 20
    else:
        ws3.cell(row=3, column=1, value="No follow-up dates set yet.").font = Font(
            name="Arial", italic=True, color="9CA3AF"
        )

    ws3.freeze_panes = "A3"

    # Return as bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# EXCEL DOWNLOAD WIDGET (reusable)
# ─────────────────────────────────────────────

def excel_download_widget(df: pd.DataFrame, label="⬇️ Download as Excel"):
    if df.empty:
        st.info("No leads to export yet.")
        return
    excel_bytes = generate_excel(df)
    filename = f"instaleads_{date.today().isoformat()}.xlsx"
    st.download_button(
        label=label,
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption("📋 3 sheets: All Leads · Summary · Follow-ups")


# ─────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────

def inject_css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&display=swap');

    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

    [data-testid="stSidebar"] { background: #0f0f0f; }
    [data-testid="stSidebar"] * { color: #f0f0f0 !important; }
    [data-testid="stSidebar"] .stRadio label {
        font-size: 15px; font-weight: 500; padding: 6px 0; cursor: pointer;
    }

    .main .block-container { max-width: 1000px; padding-top: 2rem; }

    [data-testid="metric-container"] {
        background: #f8f8f8; border: 1px solid #e8e8e8;
        border-radius: 12px; padding: 14px 18px;
    }

    .badge { display:inline-block; padding:3px 10px; border-radius:20px;
             font-size:12px; font-weight:600; letter-spacing:0.3px; }
    .badge-new            { background:#dbeafe; color:#1d4ed8; }
    .badge-interested     { background:#dcfce7; color:#166534; }
    .badge-not-interested { background:#fee2e2; color:#991b1b; }
    .badge-follow-up      { background:#fef9c3; color:#854d0e; }
    .badge-closed         { background:#f3f4f6; color:#374151; }

    .overdue-row   { background:#fff1f1; border-left:3px solid #ef4444;
                     border-radius:6px; padding:10px 14px; margin-bottom:8px; }
    .due-today-row { background:#fffbeb; border-left:3px solid #f59e0b;
                     border-radius:6px; padding:10px 14px; margin-bottom:8px; }
    .upcoming-row  { background:#f0fdf4; border-left:3px solid #22c55e;
                     border-radius:6px; padding:10px 14px; margin-bottom:8px; }

    h2 { font-weight: 600; letter-spacing: -0.5px; }
    h3 { font-weight: 500; color: #333; }
    hr { border: none; border-top: 1px solid #ececec; margin: 1.5rem 0; }

    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stSelectbox > div > div {
        border-radius: 8px !important; border-color: #e0e0e0 !important;
        font-family: 'DM Sans', sans-serif !important;
    }

    .stButton > button[kind="primary"] {
        background: #0f0f0f; color: white; border-radius: 8px;
        border: none; padding: 10px 22px; font-weight: 500;
        font-family: 'DM Sans', sans-serif;
    }
    .stButton > button[kind="primary"]:hover { background: #333; }
    .stButton > button { border-radius: 8px; font-family: 'DM Sans', sans-serif; }

    /* Download button — green */
    [data-testid="stDownloadButton"] > button {
        background: #16a34a !important; color: white !important;
        border-radius: 8px !important; border: none !important;
        font-weight: 600 !important; font-family: 'DM Sans', sans-serif !important;
    }
    [data-testid="stDownloadButton"] > button:hover {
        background: #15803d !important;
    }
    </style>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# STATUS CONFIG
# ─────────────────────────────────────────────

STATUSES    = ["New", "Interested", "Not Interested", "Follow-up", "Closed"]
STATUS_BADGE = {
    "New":            "badge-new",
    "Interested":     "badge-interested",
    "Not Interested": "badge-not-interested",
    "Follow-up":      "badge-follow-up",
    "Closed":         "badge-closed",
}
STATUS_EMOJI = {
    "New": "🔵", "Interested": "🟢", "Not Interested": "🔴",
    "Follow-up": "🟡", "Closed": "⚪",
}

def badge_html(status):
    cls = STATUS_BADGE.get(status, "badge-new")
    return f'<span class="badge {cls}">{status}</span>'


# ─────────────────────────────────────────────
# PAGE: ADD LEAD
# ─────────────────────────────────────────────

def page_add_lead():
    st.markdown("## ➕ Add New Lead")
    st.markdown("Fill in the details from the Instagram DM conversation.")
    st.markdown("<hr>", unsafe_allow_html=True)

    with st.form("add_lead_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("Customer Name *", placeholder="e.g. Aisha Mehta")
        with col2:
            username = st.text_input("Instagram Username *", placeholder="e.g. @aisha.mehta")

        col3, col4 = st.columns(2)
        with col3:
            phone = st.text_input("Phone Number", placeholder="Optional")
        with col4:
            status = st.selectbox("Lead Status", STATUSES)

        inquiry = st.text_area("What did they ask? (Inquiry)",
                               placeholder="e.g. Interested in custom hoodies, wants 5 pieces in navy blue",
                               height=100)
        notes   = st.text_area("Notes", placeholder="Any extra context or follow-up points", height=80)
        followup_date = st.date_input("Follow-up Date (optional)", value=None)

        submitted = st.form_submit_button("Save Lead", type="primary", use_container_width=True)

    if submitted:
        if not name.strip() or not username.strip():
            st.error("Customer Name and Instagram Username are required.")
        else:
            add_lead(name.strip(), username.strip(), phone.strip(),
                     inquiry.strip(), status, notes.strip(), followup_date)
            st.success(f"✅ Lead **{name}** added successfully!")
            st.balloons()

    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown("### 📥 Export Leads to Excel")
    excel_download_widget(get_all_leads())


# ─────────────────────────────────────────────
# PAGE: DASHBOARD
# ─────────────────────────────────────────────

def page_dashboard():
    st.markdown("## 📋 Lead Dashboard")

    df = get_all_leads()

    if df.empty:
        st.info("No leads yet. Add your first lead or load demo data.")
        if st.button("Load Demo Data"):
            insert_demo_data()
            st.success("Demo leads added!")
            st.rerun()
        return

    # Metrics
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Leads",  len(df))
    m2.metric("New",          len(df[df["status"] == "New"]))
    m3.metric("Interested",   len(df[df["status"] == "Interested"]))
    m4.metric("Follow-ups",   len(df[df["status"] == "Follow-up"]))

    st.markdown("<hr>", unsafe_allow_html=True)

    # Export
    excel_download_widget(df, "⬇️ Export All Leads to Excel")

    st.markdown("<hr>", unsafe_allow_html=True)

    # Filters
    fc1, fc2 = st.columns([2, 3])
    with fc1:
        status_filter = st.selectbox("Filter by Status", ["All"] + STATUSES)
    with fc2:
        search = st.text_input("Search by Name or Username", placeholder="Type to search…")

    filtered = df.copy()
    if status_filter != "All":
        filtered = filtered[filtered["status"] == status_filter]
    if search.strip():
        q = search.strip().lower()
        filtered = filtered[
            filtered["name"].str.lower().str.contains(q) |
            filtered["username"].str.lower().str.contains(q)
        ]

    st.markdown(f"**{len(filtered)} lead(s) found**")
    st.markdown("<hr>", unsafe_allow_html=True)

    if filtered.empty:
        st.warning("No leads match your filters.")
        return

    for _, row in filtered.iterrows():
        with st.expander(
            f"{STATUS_EMOJI.get(row['status'], '')}  {row['name']}  ·  "
            f"{row['username']}  ·  Added {row['date_added']}"
        ):
            c1, c2 = st.columns([3, 2])
            with c1:
                st.markdown(f"**Inquiry:** {row['inquiry'] or '—'}")
                st.markdown(f"**Notes:** {row['notes'] or '—'}")
                st.markdown(f"**Phone:** {row['phone'] or '—'}")
                st.markdown(f"**Follow-up date:** {row['followup_date'] or '—'}")
            with c2:
                st.markdown(f"**Status:** {badge_html(row['status'])}", unsafe_allow_html=True)

            st.markdown("")
            with st.form(f"update_{row['id']}"):
                new_status = st.selectbox("Update Status", STATUSES,
                                          index=STATUSES.index(row["status"]),
                                          key=f"s_{row['id']}")
                new_notes  = st.text_area("Notes", value=row["notes"] or "",
                                          key=f"n_{row['id']}", height=70)
                current_fu = datetime.strptime(row["followup_date"], "%Y-%m-%d").date() \
                             if row["followup_date"] else None
                new_fu     = st.date_input("Follow-up Date", value=current_fu,
                                           key=f"f_{row['id']}")
                uc1, uc2 = st.columns(2)
                with uc1:
                    save   = st.form_submit_button("💾 Save Changes", use_container_width=True)
                with uc2:
                    delete = st.form_submit_button("🗑 Delete Lead", use_container_width=True)

            if save:
                update_lead(row["id"], new_status, new_notes, new_fu)
                st.success("Updated!")
                st.rerun()
            if delete:
                delete_lead(row["id"])
                st.warning("Lead deleted.")
                st.rerun()

    st.markdown("<hr>", unsafe_allow_html=True)
    if st.button("Load Demo Data (adds sample leads)"):
        insert_demo_data()
        st.success("Demo leads added!")
        st.rerun()


# ─────────────────────────────────────────────
# PAGE: FOLLOW-UPS
# ─────────────────────────────────────────────

def page_followups():
    st.markdown("## 🔔 Follow-up Tracker")
    st.markdown("Leads that need your attention based on follow-up dates.")
    st.markdown("<hr>", unsafe_allow_html=True)

    df = get_all_leads()
    fu_df = df[df["followup_date"].notna() & (df["followup_date"] != "")].copy()

    if fu_df.empty:
        st.info("No follow-up dates set yet. Add one while editing a lead in the Dashboard.")
        return

    excel_download_widget(df, "⬇️ Export All Leads to Excel")
    st.markdown("<hr>", unsafe_allow_html=True)

    today = date.today()
    fu_df["followup_date_parsed"] = pd.to_datetime(fu_df["followup_date"]).dt.date
    fu_df["days_diff"] = fu_df["followup_date_parsed"].apply(lambda d: (d - today).days)
    fu_df = fu_df.sort_values("days_diff")

    overdue   = fu_df[fu_df["days_diff"] < 0]
    due_today = fu_df[fu_df["days_diff"] == 0]
    upcoming  = fu_df[fu_df["days_diff"] > 0]

    def render_card(row, css_class):
        days = int(row["days_diff"])
        if days < 0:   time_label = f"**{abs(days)} day(s) overdue**"
        elif days == 0: time_label = "**Due today**"
        else:           time_label = f"In {days} day(s)"
        st.markdown(f"""
        <div class="{css_class}">
            <strong>{row['name']}</strong> &nbsp;
            <span style="color:#888">{row['username']}</span><br>
            <span style="font-size:13px; color:#555">
                📅 {row['followup_date']} &nbsp;|&nbsp; {time_label} &nbsp;|&nbsp;
                {badge_html(row['status'])}
            </span><br>
            <span style="font-size:13px; margin-top:4px; display:block; color:#444">
                💬 {row['inquiry'] or '—'}
            </span>
            {f'<span style="font-size:12px; color:#666; margin-top:4px; display:block;">📝 {row["notes"]}</span>'
              if row["notes"] else ''}
        </div>
        """, unsafe_allow_html=True)

    if not overdue.empty:
        st.markdown(f"### 🔴 Overdue ({len(overdue)})")
        for _, row in overdue.iterrows(): render_card(row, "overdue-row")

    if not due_today.empty:
        st.markdown(f"### 🟡 Due Today ({len(due_today)})")
        for _, row in due_today.iterrows(): render_card(row, "due-today-row")

    if not upcoming.empty:
        st.markdown(f"### 🟢 Upcoming ({len(upcoming)})")
        for _, row in upcoming.iterrows(): render_card(row, "upcoming-row")

    if overdue.empty and due_today.empty and upcoming.empty:
        st.success("You're all caught up! No pending follow-ups.")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="InstaLeads — DM CRM",
        page_icon="📩",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    inject_css()
    init_db()

    with st.sidebar:
        st.markdown("## 📩 InstaLeads")
        st.markdown("*Your Instagram DM CRM*")
        st.markdown("---")

        page = st.radio(
            "Navigate",
            ["➕ Add Lead", "📋 Dashboard", "🔔 Follow-ups"],
            label_visibility="collapsed",
        )

        st.markdown("---")
        df = get_all_leads()
        if not df.empty:
            st.markdown("**Quick Stats**")
            for s in STATUSES:
                cnt = len(df[df["status"] == s])
                if cnt > 0:
                    st.markdown(f"{STATUS_EMOJI[s]} {s}: **{cnt}**")
            st.markdown("---")
            st.markdown("**Export**")
            excel_bytes = generate_excel(df)
            st.download_button(
                label="⬇️ Download Excel",
                data=excel_bytes,
                file_name=f"instaleads_{date.today().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.markdown("---")
        st.markdown(
            "<span style='font-size:11px; color:#888'>Built for small business owners.<br>"
            "Track every DM. Miss no customer.</span>",
            unsafe_allow_html=True,
        )

    if page == "➕ Add Lead":
        page_add_lead()
    elif page == "📋 Dashboard":
        page_dashboard()
    elif page == "🔔 Follow-ups":
        page_followups()


if __name__ == "__main__":
    main()
